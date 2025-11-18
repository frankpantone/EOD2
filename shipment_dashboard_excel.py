import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from datetime import datetime
import sys
import glob
import os

# Find all CSV files in the current directory
csv_files = glob.glob("*.csv")

if len(csv_files) == 0:
    print("[ERROR] No CSV files found in the current directory.")
    sys.exit(1)

# Separate main EOD file from EOD Update-2 file
eod_update2_file = None
main_csv_files = []

for f in csv_files:
    if 'EOD Update-2' in f or 'EOD Update_2' in f:
        eod_update2_file = f
    else:
        main_csv_files.append(f)

if len(main_csv_files) == 1:
    csv_file = main_csv_files[0]
    print(f"[OK] Found main CSV file: {csv_file}")
elif len(main_csv_files) > 1:
    # Multiple CSV files found - use the most recently modified one
    csv_file = max(main_csv_files, key=os.path.getmtime)
    print(f"[OK] Multiple CSV files found. Using most recent: {csv_file}")
    print(f"[INFO] Other files in directory: {', '.join([f for f in main_csv_files if f != csv_file])}")
else:
    print("[ERROR] No main CSV file found.")
    sys.exit(1)

if eod_update2_file:
    print(f"[OK] Found EOD Update-2 file: {eod_update2_file}")

# Read the CSV file
try:
    df = pd.read_csv(csv_file)
    print(f"[OK] Loaded {len(df)} records from {csv_file}")
except Exception as e:
    print(f"[ERROR] Failed to read file: {e}")
    sys.exit(1)

# Convert Created Date to datetime
df['Created Date'] = pd.to_datetime(df['Created Date'])

# Rule: Remove orders with tag "CSRM, Quote"
initial_count = len(df)
df = df[~df['Tags'].str.contains('Quote', case=False, na=False)]
filtered_count = len(df)
print(f"[OK] Filtered out {initial_count - filtered_count} records with 'Quote' tag")
print(f"[OK] Working with {filtered_count} records")

# Get today's date (from the data)
today = df['Created Date'].max().date()
print(f"[OK] Latest date in data: {today}")

# Filter for today's shipments
df_today = df[df['Created Date'].dt.date == today]
total_today = len(df_today)

# Calculate total shipments (all dates)
total_all = len(df)

# Calculate increase
increase = total_today
increase_pct = (increase / total_all * 100) if total_all > 0 else 0

# Most shipped vehicle
most_shipped_vehicle = df['Vehicle Info'].value_counts().head(1)
most_shipped_vehicle_name = most_shipped_vehicle.index[0] if len(most_shipped_vehicle) > 0 else "N/A"
most_shipped_vehicle_count = most_shipped_vehicle.values[0] if len(most_shipped_vehicle) > 0 else 0

# Weighted average distance
df['Distance'] = pd.to_numeric(df['Distance'], errors='coerce')
weighted_avg_distance = df['Distance'].mean()

# Create pivot table: Rows = Customer Business Name, Columns = Tags, Values = Count of VIN #
pivot_table = pd.pivot_table(
    df,
    values='VIN #',
    index='Customer Business Name',
    columns='Tags',
    aggfunc='count',
    fill_value=0
)

# Sort by total shipments per customer
pivot_table['Total'] = pivot_table.sum(axis=1)
pivot_table = pivot_table.sort_values('Total', ascending=False)

print(f"[OK] Pivot table created with {len(pivot_table)} customers and {len(pivot_table.columns)-1} tag types")

# Create pivot table for TODAY's shipments only
pivot_table_today = pd.pivot_table(
    df_today,
    values='VIN #',
    index='Customer Business Name',
    columns='Tags',
    aggfunc='count',
    fill_value=0
)

# Sort by total shipments per customer
pivot_table_today['Total'] = pivot_table_today.sum(axis=1)
pivot_table_today = pivot_table_today.sort_values('Total', ascending=False)

print(f"[OK] Today's pivot table created with {len(pivot_table_today)} customers")

# Tag distribution
tag_distribution = df.groupby('Tags').size().reset_index(name='Count')
tag_distribution = tag_distribution.sort_values('Count', ascending=False)

# Top vehicles
top_vehicles = df['Vehicle Info'].value_counts().head(10).reset_index()
top_vehicles.columns = ['Vehicle', 'Count']

# Process EOD Update-2 file for CarMax unique VINs with New status and no tags
carmax_vins_by_date = pd.DataFrame()
carmax_unique_vins_total = 0

if eod_update2_file:
    try:
        df_update2 = pd.read_csv(eod_update2_file)
        print(f"[OK] Loaded {len(df_update2)} records from EOD Update-2 file")
        
        # Filter for CarMax, New status, and no tags (empty/null/whitespace)
        carmax_new_no_tags = df_update2[
            (df_update2['Customer Business Name'].str.contains('CarMax', case=False, na=False)) & 
            (df_update2['Vehicle Status'].str.contains('New', case=False, na=False)) &
            (df_update2['Tags'].isna() | (df_update2['Tags'].str.strip() == ''))
        ]
        
        if len(carmax_new_no_tags) > 0:
            # Convert Created Date to datetime
            carmax_new_no_tags = carmax_new_no_tags.copy()
            carmax_new_no_tags['Created Date'] = pd.to_datetime(carmax_new_no_tags['Created Date'])
            
            # Group by Created Date and count unique VINs
            carmax_vins_by_date = carmax_new_no_tags.groupby(
                carmax_new_no_tags['Created Date'].dt.date
            )['VIN #'].nunique().reset_index()
            carmax_vins_by_date.columns = ['Created Date', 'Unique VINs']
            carmax_vins_by_date = carmax_vins_by_date.sort_values('Created Date')
            
            # Calculate total unique VINs
            carmax_unique_vins_total = carmax_new_no_tags['VIN #'].nunique()
            
            print(f"[OK] Found {carmax_unique_vins_total} unique CarMax VINs across {len(carmax_vins_by_date)} dates")
    except Exception as e:
        print(f"[WARNING] Could not process EOD Update-2 file: {e}")
        carmax_unique_vins_total = 0

# Create Excel file
output_file = f"shipment_dashboard_{today.strftime('%Y-%m-%d')}.xlsx"

# Create a Pandas Excel writer using openpyxl as the engine
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    
    # SHEET 1: Dashboard Summary
    wb = writer.book
    ws_summary = wb.create_sheet('Dashboard Summary', 0)
    
    # Title
    ws_summary['A1'] = 'SHIPMENT DASHBOARD'
    ws_summary['A1'].font = Font(size=24, bold=True, color='2c3e50')
    ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_summary.merge_cells('A1:F1')
    ws_summary.row_dimensions[1].height = 35
    
    # Report Date
    ws_summary['A2'] = f'Report Date: {today.strftime("%B %d, %Y")}'
    ws_summary['A2'].font = Font(size=12, color='7f8c8d')
    ws_summary['A2'].alignment = Alignment(horizontal='center')
    ws_summary.merge_cells('A2:F2')
    
    # Key Metrics Headers
    row = 4
    metrics_data = [
        ('SHIPMENTS CREATED TODAY', total_today, f'Date: {today}', '667eea'),
        ('TODAY VS TOTAL', increase, f'{increase_pct:.1f}% of total ({total_all} total)', 'f5576c'),
        ('MOST SHIPPED VEHICLE', most_shipped_vehicle_count, most_shipped_vehicle_name, '00f2fe'),
        ('AVERAGE DISTANCE', f'{weighted_avg_distance:.0f}', 'miles per shipment', '38f9d7')
    ]
    
    for idx, (label, value, subtitle, color) in enumerate(metrics_data):
        col_offset = (idx % 2) * 3 + 1
        row_offset = (idx // 2) * 4 + row
        
        # Label
        cell = ws_summary.cell(row=row_offset, column=col_offset)
        cell.value = label
        cell.font = Font(size=10, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws_summary.merge_cells(start_row=row_offset, start_column=col_offset, 
                              end_row=row_offset, end_column=col_offset+1)
        
        # Value
        cell = ws_summary.cell(row=row_offset+1, column=col_offset)
        cell.value = value
        cell.font = Font(size=28, bold=True, color='2c3e50')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws_summary.merge_cells(start_row=row_offset+1, start_column=col_offset, 
                              end_row=row_offset+1, end_column=col_offset+1)
        ws_summary.row_dimensions[row_offset+1].height = 40
        
        # Subtitle
        cell = ws_summary.cell(row=row_offset+2, column=col_offset)
        cell.value = subtitle
        cell.font = Font(size=9, color='7f8c8d')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws_summary.merge_cells(start_row=row_offset+2, start_column=col_offset, 
                              end_row=row_offset+2, end_column=col_offset+1)
    
    # Summary Information
    summary_row = 16
    ws_summary[f'A{summary_row}'] = 'DATA SUMMARY'
    ws_summary[f'A{summary_row}'].font = Font(size=14, bold=True, color='2c3e50')
    ws_summary.merge_cells(f'A{summary_row}:F{summary_row}')
    
    summary_info = [
        ['Total Records Processed:', filtered_count],
        ['Filtered Out (Quote tags):', initial_count - filtered_count],
        ['Date Range:', f"{df['Created Date'].min().strftime('%m/%d/%Y')} to {df['Created Date'].max().strftime('%m/%d/%Y')}"],
        ['Number of Customers:', len(pivot_table)],
        ['Number of Tag Types:', len(pivot_table.columns)-1]
    ]
    
    for idx, (label, value) in enumerate(summary_info):
        row_num = summary_row + idx + 1
        ws_summary[f'A{row_num}'] = label
        ws_summary[f'A{row_num}'].font = Font(bold=True)
        ws_summary[f'B{row_num}'] = value
    
    # Column widths
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 5
    ws_summary.column_dimensions['D'].width = 30
    ws_summary.column_dimensions['E'].width = 20
    
    # SHEET 2: Pivot Table
    ws_pivot = wb.create_sheet('Pivot Table')
    
    # Define formatting
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    total_fill = PatternFill(start_color='ffd700', end_color='ffd700', fill_type='solid')
    total_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # TABLE 1: ALL SHIPMENTS
    current_row = 1
    
    # Get date range
    first_date = df['Created Date'].min().strftime('%m/%d/%Y')
    last_date = df['Created Date'].max().strftime('%m/%d/%Y')
    
    # Title with date range and total count
    ws_pivot.cell(row=current_row, column=1).value = f'Count of VIN by Customer and Tag Type ({first_date} - {last_date}) - {total_all}'
    ws_pivot.cell(row=current_row, column=1).font = Font(size=16, bold=True, color='2c3e50')
    ws_pivot.merge_cells(start_row=current_row, start_column=1, 
                        end_row=current_row, end_column=len(pivot_table.columns)+1)
    current_row += 2
    
    # Headers
    pivot_reset = pivot_table.reset_index()
    headers = list(pivot_reset.columns)
    for col_num, header in enumerate(headers, start=1):
        cell = ws_pivot.cell(row=current_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    ws_pivot.row_dimensions[current_row].height = 30  # Set header row height
    
    current_row += 1
    
    # Data rows
    for idx, row_data in pivot_reset.iterrows():
        for col_num, value in enumerate(row_data, start=1):
            cell = ws_pivot.cell(row=current_row, column=col_num)
            cell.value = value if not isinstance(value, (int, float)) or value > 0 else ''
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Alternate row colors
            if current_row % 2 == 0:
                cell.fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
            
            # Highlight Total column
            if col_num == len(headers):
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='e9ecef', end_color='e9ecef', fill_type='solid')
        
        current_row += 1
    
    # TOTALS ROW for Table 1
    cell = ws_pivot.cell(row=current_row, column=1)
    cell.value = 'TOTAL'
    cell.fill = total_fill
    cell.font = total_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    
    for col_num in range(2, len(headers) + 1):
        col_name = headers[col_num - 1]
        if col_name == 'Customer Business Name':
            continue
        total_value = pivot_table[col_name].sum()
        cell = ws_pivot.cell(row=current_row, column=col_num)
        cell.value = int(total_value)
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # SMALL TABLE: CarMax Unique VINs by Created Date (to the right of main table)
    carmax_table_start_col = len(headers) + 2  # Start 1 column after the main table
    carmax_table_row = 1  # Start at the top
    
    # Title for CarMax table
    cell = ws_pivot.cell(row=carmax_table_row, column=carmax_table_start_col)
    cell.value = 'CarMax VINs - New Status (No Tags)'
    cell.font = Font(size=14, bold=True, color='2c3e50')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_pivot.merge_cells(start_row=carmax_table_row, start_column=carmax_table_start_col,
                        end_row=carmax_table_row, end_column=carmax_table_start_col + 1)
    carmax_table_row += 2
    
    # Headers for CarMax table
    carmax_headers = ['Created Date', 'Unique VINs']
    for col_offset, header in enumerate(carmax_headers):
        cell = ws_pivot.cell(row=carmax_table_row, column=carmax_table_start_col + col_offset)
        cell.value = header
        cell.fill = PatternFill(start_color='f5576c', end_color='f5576c', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    carmax_table_row += 1
    
    # Display unique VINs by date
    if len(carmax_vins_by_date) > 0:
        for idx, row_data in carmax_vins_by_date.iterrows():
            # Date column
            cell = ws_pivot.cell(row=carmax_table_row, column=carmax_table_start_col)
            cell.value = row_data['Created Date'].strftime('%m/%d/%Y')
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Unique VINs column
            cell = ws_pivot.cell(row=carmax_table_row, column=carmax_table_start_col + 1)
            cell.value = int(row_data['Unique VINs'])
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True, color='f5576c')
            
            if carmax_table_row % 2 == 0:
                ws_pivot.cell(row=carmax_table_row, column=carmax_table_start_col).fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
                ws_pivot.cell(row=carmax_table_row, column=carmax_table_start_col + 1).fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
            
            carmax_table_row += 1
        
        # Total row for CarMax table
        cell = ws_pivot.cell(row=carmax_table_row, column=carmax_table_start_col)
        cell.value = 'TOTAL'
        cell.fill = PatternFill(start_color='ffa502', end_color='ffa502', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        cell = ws_pivot.cell(row=carmax_table_row, column=carmax_table_start_col + 1)
        cell.value = carmax_unique_vins_total
        cell.fill = PatternFill(start_color='ffa502', end_color='ffa502', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    else:
        # No CarMax data
        cell = ws_pivot.cell(row=carmax_table_row, column=carmax_table_start_col)
        cell.value = 'No data found'
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws_pivot.merge_cells(start_row=carmax_table_row, start_column=carmax_table_start_col,
                            end_row=carmax_table_row, end_column=carmax_table_start_col + 1)
    
    # Set column widths for CarMax table
    ws_pivot.column_dimensions[chr(64 + carmax_table_start_col)].width = 25
    ws_pivot.column_dimensions[chr(64 + carmax_table_start_col + 1)].width = 20
    
    current_row += 3  # Add spacing
    
    # TABLE 2: TODAY'S SHIPMENTS
    if len(pivot_table_today) > 0:
        # Calculate overall percentage
        overall_percentage = (total_today / total_all * 100) if total_all > 0 else 0
        
        # Title with today's date, count, and percentage
        ws_pivot.cell(row=current_row, column=1).value = f'Count of VIN Created Today ({today.strftime("%m/%d/%Y")}) - {total_today} ({overall_percentage:.1f}% Increase)'
        ws_pivot.cell(row=current_row, column=1).font = Font(size=16, bold=True, color='2c3e50')
        ws_pivot.merge_cells(start_row=current_row, start_column=1, 
                            end_row=current_row, end_column=len(pivot_table_today.columns)+2)
        current_row += 2
        
        # Headers - add % Increase column
        pivot_today_reset = pivot_table_today.reset_index()
        headers_today = list(pivot_today_reset.columns) + ['% Increase']
        for col_num, header in enumerate(headers_today, start=1):
            cell = ws_pivot.cell(row=current_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        ws_pivot.row_dimensions[current_row].height = 30  # Set header row height
        
        current_row += 1
        
        # Data rows
        for idx, row_data in pivot_today_reset.iterrows():
            customer_name = row_data['Customer Business Name']
            
            for col_num, value in enumerate(row_data, start=1):
                cell = ws_pivot.cell(row=current_row, column=col_num)
                cell.value = value if not isinstance(value, (int, float)) or value > 0 else ''
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Alternate row colors
                if current_row % 2 == 0:
                    cell.fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
                
                # Highlight Total column
                if col_num == len(pivot_today_reset.columns):
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='e9ecef', end_color='e9ecef', fill_type='solid')
            
            # Add percentage column
            today_total = row_data['Total']
            if customer_name in pivot_table.index:
                all_time_total = pivot_table.loc[customer_name, 'Total']
                percentage = (today_total / all_time_total) if all_time_total > 0 else 0
            else:
                percentage = 0
            
            cell = ws_pivot.cell(row=current_row, column=len(headers_today))
            cell.value = percentage  # Store as numeric value (0.270 for 27%)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            cell.font = Font(color='0066cc', bold=True, size=11)
            cell.number_format = '0.0%'  # Format as percentage with 1 decimal
            
            if current_row % 2 == 0:
                cell.fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
            else:
                cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            
            current_row += 1
        
        # TOTALS ROW for Table 2
        cell = ws_pivot.cell(row=current_row, column=1)
        cell.value = 'TOTAL'
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        for col_num in range(2, len(pivot_today_reset.columns) + 1):
            col_name = pivot_today_reset.columns[col_num - 1]
            if col_name == 'Customer Business Name':
                continue
            total_value = pivot_table_today[col_name].sum()
            cell = ws_pivot.cell(row=current_row, column=col_num)
            cell.value = int(total_value)
            cell.fill = total_fill
            cell.font = total_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Add overall percentage in the % Increase column
        cell = ws_pivot.cell(row=current_row, column=len(headers_today))
        cell.value = overall_percentage / 100  # Store as numeric value (0.135 for 13.5%)
        cell.fill = total_fill
        cell.font = total_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        cell.border = thin_border
        cell.number_format = '0.0%'  # Format as percentage with 1 decimal
    
    # Adjust column widths - set based on header content
    ws_pivot.column_dimensions['A'].width = 45  # Customer Business Name (increased for long names)
    
    # Set width for each column in Table 1 based on header length
    for col_num, header in enumerate(headers, start=1):
        col_letter = chr(64 + col_num) if col_num <= 26 else chr(64 + col_num // 26) + chr(64 + col_num % 26)
        if col_num == 1:
            ws_pivot.column_dimensions[col_letter].width = 45  # Customer name (increased for long names)
        elif 'Total' in str(header):
            ws_pivot.column_dimensions[col_letter].width = 12
        else:
            # Adjust based on header text length
            header_length = len(str(header))
            ws_pivot.column_dimensions[col_letter].width = max(header_length + 3, 22)
    
    # Set width for % Increase column if today's table exists
    if len(pivot_table_today) > 0:
        increase_col = len(headers_today)
        col_letter = chr(64 + increase_col) if increase_col <= 26 else chr(64 + increase_col // 26) + chr(64 + increase_col % 26)
        ws_pivot.column_dimensions[col_letter].width = 15
    
    # SHEET 3: Tag Distribution
    tag_distribution.to_excel(writer, sheet_name='Tag Distribution', index=False, startrow=2)
    
    ws_tags = writer.sheets['Tag Distribution']
    
    # Title
    ws_tags['A1'] = 'Shipment Distribution by Tag Type'
    ws_tags['A1'].font = Font(size=16, bold=True, color='2c3e50')
    ws_tags.merge_cells('A1:B1')
    
    # Format header
    for col_num in range(1, 3):
        cell = ws_tags.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Format data
    for row_num in range(4, len(tag_distribution) + 4):
        for col_num in range(1, 3):
            cell = ws_tags.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
    
    ws_tags.column_dimensions['A'].width = 30
    ws_tags.column_dimensions['B'].width = 15
    
    # Add Pie Chart
    pie = PieChart()
    labels = Reference(ws_tags, min_col=1, min_row=4, max_row=len(tag_distribution) + 3)
    data = Reference(ws_tags, min_col=2, min_row=3, max_row=len(tag_distribution) + 3)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Tag Distribution"
    pie.width = 15
    pie.height = 10
    ws_tags.add_chart(pie, "D3")
    
    # SHEET 4: Top Vehicles
    top_vehicles.to_excel(writer, sheet_name='Top Vehicles', index=False, startrow=2)
    
    ws_vehicles = writer.sheets['Top Vehicles']
    
    # Title
    ws_vehicles['A1'] = 'Top 10 Most Shipped Vehicles'
    ws_vehicles['A1'].font = Font(size=16, bold=True, color='2c3e50')
    ws_vehicles.merge_cells('A1:B1')
    
    # Format header
    for col_num in range(1, 3):
        cell = ws_vehicles.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Format data
    for row_num in range(4, len(top_vehicles) + 4):
        for col_num in range(1, 3):
            cell = ws_vehicles.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
    
    ws_vehicles.column_dimensions['A'].width = 35
    ws_vehicles.column_dimensions['B'].width = 15
    
    # Add Bar Chart
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "Top Vehicles"
    bar_chart.y_axis.title = 'Count'
    bar_chart.x_axis.title = 'Vehicle'
    
    data = Reference(ws_vehicles, min_col=2, min_row=3, max_row=len(top_vehicles) + 3)
    cats = Reference(ws_vehicles, min_col=1, min_row=4, max_row=len(top_vehicles) + 3)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)
    bar_chart.width = 15
    bar_chart.height = 10
    ws_vehicles.add_chart(bar_chart, "D3")
    
    # SHEET 5: Raw Data (Filtered)
    df_export = df.copy()
    df_export['Created Date'] = df_export['Created Date'].dt.strftime('%m/%d/%Y')
    df_export.to_excel(writer, sheet_name='Raw Data', index=False, startrow=1)
    
    ws_raw = writer.sheets['Raw Data']
    
    # Title
    ws_raw['A1'] = 'Filtered Shipment Data (Quotes Removed)'
    ws_raw['A1'].font = Font(size=14, bold=True, color='2c3e50')
    ws_raw.merge_cells('A1:W1')
    
    # Format header
    for col_num in range(1, len(df_export.columns) + 1):
        cell = ws_raw.cell(row=2, column=col_num)
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Auto-adjust column widths
    for col_idx, column in enumerate(ws_raw.iter_cols(min_row=2, max_row=len(df_export)+2), start=1):
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_raw.column_dimensions[column_letter].width = adjusted_width
    
    # Remove default sheet if it exists
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

print(f"\n[SUCCESS] Excel Dashboard created successfully: {output_file}")
print(f"\n[METRICS] Key Metrics:")
print(f"   - Shipments Created Today: {total_today}")
print(f"   - Total Shipments (All Time): {total_all}")
print(f"   - Today's Percentage: {increase_pct:.1f}%")
print(f"   - Most Shipped Vehicle: {most_shipped_vehicle_name} ({most_shipped_vehicle_count} units)")
print(f"   - Average Distance: {weighted_avg_distance:.2f} miles")
if carmax_unique_vins_total > 0:
    print(f"   - CarMax Unique VINs (New, No Tags): {carmax_unique_vins_total}")
print(f"\n[SHEETS] Excel file contains 5 sheets:")
print(f"   1. Dashboard Summary - Key metrics and overview")
print(f"   2. Pivot Table - Customers x Tags breakdown + CarMax Unique VINs by Date")
print(f"   3. Tag Distribution - Shipments by tag type (with chart)")
print(f"   4. Top Vehicles - Most shipped vehicles (with chart)")
print(f"   5. Raw Data - Complete filtered dataset")
print(f"\n[INFO] Open the Excel file to view your interactive dashboard!")

